from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseForbidden, JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib.auth.models import User
from django.contrib import messages
from io import BytesIO
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime

from .models import Etudiant, Note, UE, Departement, Filiere, Niveau
from .forms import EtudiantForm, TeacherCreationForm
import json


def home(request):
    """Homepage with quick stats and recent notes."""
    stats = {
        'departements': Departement.objects.count(),
        'filieres': Filiere.objects.count(),
        'niveaux': Niveau.objects.count(),
        'ues': UE.objects.count(),
        'etudiants': Etudiant.objects.count(),
    }
    recent_notes = Note.objects.select_related('etudiant', 'ue').order_by('-id')[:5]
    return render(request, 'pages/home_adminlte.html', {'stats': stats, 'recent_notes': recent_notes})


from django.core.paginator import Paginator, EmptyPage
from django.db.models import OuterRef, Subquery, Value, FloatField, ExpressionWrapper, F
from django.db.models.functions import Coalesce


def etudiant_list(request):
    # Cascade filters: departement -> filiere -> niveau -> optional ue
    deps = Departement.objects.order_by('id')
    dep_id = request.GET.get('departement')
    dep = None
    if dep_id:
        try:
            dep = Departement.objects.get(pk=dep_id)
        except Departement.DoesNotExist:
            dep = deps.first()
    else:
        dep = deps.first()

    filieres = Filiere.objects.filter(departement=dep).order_by('id') if dep else Filiere.objects.order_by('id')
    fil_id = request.GET.get('filiere')
    fil = None
    if fil_id:
        fil = filieres.filter(pk=fil_id).first() or filieres.first()
    else:
        fil = filieres.first()

    # choose niveaux relevant to the filiere when possible (students attached to filiere)
    if fil:
        niveaux_qs = Niveau.objects.filter(etudiant__filiere=fil).distinct().order_by('id')
    else:
        niveaux_qs = Niveau.objects.order_by('id')
    niv_id = request.GET.get('niveau')
    niv = None
    if niv_id:
        niv = niveaux_qs.filter(pk=niv_id).first() or niveaux_qs.first()
    else:
        niv = niveaux_qs.first()

    # Semester filter (1 or 2)
    semester = request.GET.get('semester', 1)
    try:
        semester = int(semester)
        if semester not in (1, 2):
            semester = 1
    except (ValueError, TypeError):
        semester = 1

    # UEs within filiere+niveau+semester (for optional column)
    ues = UE.objects.filter(filiere=fil, niveau=niv, semester=semester).order_by('code') if fil and niv else UE.objects.none()
    ue_id = request.GET.get('ue')
    ue_selected = None
    if ue_id:
        ue_selected = ues.filter(pk=ue_id).first()

    # student queryset limited by filiere + niveau to avoid loading entire DB
    students_qs = Etudiant.objects.none()
    if fil and niv:
        sort = request.GET.get('sort', 'nom')
        # allow 'note' sorting only when UE selected
        if sort not in ('nom', 'matricule', 'note'):
            sort = 'nom'
        students_qs = Etudiant.objects.filter(filiere=fil, niveau=niv)

        # if sorting by note and a UE is selected, annotate each student with the UE final and sort by it
        if sort == 'note' and ue_selected:
            # compute final = cc * cc_weight/100 + tp * tp_weight/100 + sn * sn_weight/100
            final_expr = ExpressionWrapper(
                F('cc') * ue_selected.cc_weight / 100.0 +
                F('tp') * ue_selected.tp_weight / 100.0 +
                F('sn') * ue_selected.sn_weight / 100.0,
                output_field=FloatField()
            )
            note_subq = Note.objects.filter(etudiant=OuterRef('pk'), ue=ue_selected).annotate(final_calc=final_expr).values('final_calc')[:1]
            students_qs = students_qs.annotate(note_final=Coalesce(Subquery(note_subq, output_field=FloatField()), Value(-1.0))).order_by('-note_final', 'nom')
        else:
            # regular ordering by field
            students_qs = students_qs.order_by(sort)

    # pagination (default page_size 20)
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    paginator = Paginator(students_qs, page_size)
    try:
        students_page = paginator.page(page)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages) if paginator.num_pages else []

    total_students = students_qs.count()

    # if UE selected, fetch notes for the students on that page
    notes_map = {}
    if ue_selected:
        notes = Note.objects.filter(ue=ue_selected, etudiant__in=students_page.object_list).select_related('etudiant')
        notes_map = {n.etudiant_id: n for n in notes}

    # assemble rows so template lookup is straightforward
    rows = []
    for s in students_page.object_list:
        n = notes_map.get(s.id)
        rows.append({'etudiant': s, 'note_final': (n.final if n and n.final is not None else None)})

    # build base query for pagination links (preserve filters but not 'page')
    base_qs = request.GET.copy()
    if 'page' in base_qs:
        base_qs.pop('page')
    base_query = base_qs.urlencode()

    context = {
        'departements': deps,
        'filieres': filieres,
        'niveaux': niveaux_qs,
        'ues': ues,
        'selected_departement': getattr(dep, 'id', None),
        'selected_filiere': getattr(fil, 'id', None),
        'selected_niveau': getattr(niv, 'id', None),
        'selected_semester': semester,
        'selected_ue': getattr(ue_selected, 'id', None),
        'students_page': students_page,
        'rows': rows,
        'page': page,
        'page_size': page_size,
        'total_students': total_students,
        'base_query': base_query,
        'sort': request.GET.get('sort', 'nom'),
        'current_path': request.get_full_path(),
    }
    return render(request, 'pages/etudiants_list_adminlte.html', context)


@login_required
def etudiant_create(request):
    if not request.user.is_staff:
        return HttpResponseForbidden()
    if request.method == 'POST':
        form = EtudiantForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('etudiants_list')
    else:
        form = EtudiantForm()
    return render(request, 'pages/etudiant_form_adminlte.html', {'form': form})


def moyenne_etudiant(request, etudiant_id):
    etudiant = get_object_or_404(Etudiant, id=etudiant_id)
    notes = Note.objects.filter(etudiant=etudiant).select_related('ue')

    # compute final per UE and weighted moyenne by UE.credit (ignore UEs with missing final)
    notes_with_final = [n for n in notes if n.final is not None]
    moyenne = None
    if notes_with_final:
        sum_weighted = sum(n.final * n.ue.credit for n in notes_with_final)
        sum_credits = sum(n.ue.credit for n in notes_with_final)
        if sum_credits > 0:
            moyenne = round(sum_weighted / sum_credits, 2)

    # preserve optional 'next' param so template can return to filtered list
    return render(request, 'pages/moyenne_adminlte.html', {'etudiant': etudiant, 'notes': notes, 'average': moyenne, 'next': request.GET.get('next', '/')})


def moyenne_etudiant_pdf(request, etudiant_id):
    """Export student notes transcript as PDF."""
    etudiant = get_object_or_404(Etudiant, id=etudiant_id)
    notes = Note.objects.filter(etudiant=etudiant).select_related('ue').order_by('ue__code')

    # compute final per UE and weighted moyenne
    notes_with_final = [n for n in notes if n.final is not None]
    moyenne = None
    if notes_with_final:
        sum_weighted = sum(n.final * n.ue.credit for n in notes_with_final)
        sum_credits = sum(n.ue.credit for n in notes_with_final)
        if sum_credits > 0:
            moyenne = round(sum_weighted / sum_credits, 2)

    # Create PDF content
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20, bottomMargin=20)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph(f"Relevé de Notes - {etudiant.nom}", title_style)
    elements.append(title)
    
    # Student info
    info_style = styles['Normal']
    info_text = f"<b>Matricule:</b> {etudiant.matricule} | <b>Filière:</b> {etudiant.filiere.nom} | <b>Niveau:</b> {etudiant.niveau.nom}"
    elements.append(Paragraph(info_text, info_style))
    elements.append(Spacer(1, 12))
    
    # Notes table
    table_data = [
        ['UE', 'Code', 'Crédit', 'CC', 'TP', 'SN', 'Final', 'État'],
    ]
    
    for note in notes:
        table_data.append([
            note.ue.nom,
            note.ue.code,
            str(note.ue.credit),
            str(note.cc) if note.cc is not None else '—',
            str(note.tp) if note.tp is not None else '—',
            str(note.sn) if note.sn is not None else '—',
            f"{note.final:.2f}" if note.final is not None else '—',
            'Éliminé' if note.is_eliminated else 'Valide',
        ])
    
    table = Table(table_data, colWidths=[2.2*inch, 0.8*inch, 0.7*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.7*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 12))
    
    # Moyenne
    if moyenne is not None:
        moyenne_text = f"<b>Moyenne pondérée:</b> {moyenne:.2f}"
    else:
        moyenne_text = "<b>Moyenne pondérée:</b> Non calculable (notes incomplètes)"
    elements.append(Paragraph(moyenne_text, info_style))
    
    elements.append(Spacer(1, 12))
    date_text = f"<i>Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}</i>"
    elements.append(Paragraph(date_text, info_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Return as attachment
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="releve_notes_{etudiant.matricule}.pdf"'
    return response


def logout_view(request):
    """Log out the user and redirect to students list with a message."""
    if request.user.is_authenticated:
        logout(request)
        messages.info(request, 'Vous êtes maintenant déconnecté.')
    return redirect('etudiants_list')


# ---------- Tableau & API ----------
@login_required
def tableau_notes(request):
    if not request.user.is_staff:
        return HttpResponseForbidden()
    deps = Departement.objects.all()
    filieres = Filiere.objects.all()
    niveaux = Niveau.objects.all()
    return render(request, 'pages/tableau_notes_adminlte.html', {'departements': deps, 'filieres': filieres, 'niveaux': niveaux})


@login_required
def notes_json(request):
    # returns UEs and students with notes according to filters, with optional pagination
    dep_id = request.GET.get('departement')
    fil_id = request.GET.get('filiere')
    niv_id = request.GET.get('niveau')

    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 25))

    ues_qs = UE.objects.all()
    if dep_id:
        ues_qs = ues_qs.filter(filiere__departement_id=dep_id)
    if fil_id:
        ues_qs = ues_qs.filter(filiere_id=fil_id)
    if niv_id:
        ues_qs = ues_qs.filter(niveau_id=niv_id)

    # helper to check management rights
    def user_manages_ue(user, ue):
        return user.is_superuser or user.is_staff or ue.instructors.filter(pk=user.pk).exists() or (hasattr(user, 'ues') and user.ues.filter(pk=ue.pk).exists())

    # mark UEs editable for the current user
    ues_list = []
    for u in ues_qs.order_by('code'):
        can_edit = user_manages_ue(request.user, u)
        ues_list.append({'id': u.id, 'code': u.code, 'nom': u.nom, 'credit': u.credit, 'editable': can_edit})

    students_qs = Etudiant.objects.all()
    if dep_id:
        students_qs = students_qs.filter(filiere__departement_id=dep_id)
    if fil_id:
        students_qs = students_qs.filter(filiere_id=fil_id)
    if niv_id:
        students_qs = students_qs.filter(niveau_id=niv_id)

    total_students = students_qs.count()
    # ordering by nom
    students_qs = students_qs.order_by('nom')

    # pagination
    start = (page - 1) * page_size
    end = start + page_size
    students_page = students_qs[start:end]

    students = []
    u_ids = [u['id'] for u in ues_list]
    for s in students_page:
        row = {'id': s.id, 'nom': s.nom, 'matricule': s.matricule, 'notes': {}}
        notes = Note.objects.filter(etudiant=s, ue__in=u_ids).select_related('ue')
        notes_map = {n.ue.id: n for n in notes}
        for u in ues_list:
            n = notes_map.get(u['id'])
            if n:
                row['notes'][str(u['id'])] = {
                    'cc': n.cc,
                    'tp': n.tp,
                    'sn': n.sn,
                    'final': n.final,
                    'is_eliminated': n.is_eliminated,
                    'note_id': n.id,
                }
            else:
                row['notes'][str(u['id'])] = None
        students.append(row)

    return JsonResponse({'ues': ues_list, 'students': students, 'page': page, 'page_size': page_size, 'total_students': total_students})


@login_required
@require_POST
def note_update(request, note_id):
    try:
        data = json.loads(request.body.decode())
    except Exception:
        return HttpResponseBadRequest('Invalid JSON')

    cc = data.get('cc')
    tp = data.get('tp')
    sn = data.get('sn')

    # basic validation: None or 0-20
    def valid_val(x):
        return x is None or (isinstance(x, (int, float)) and 0 <= x <= 20)

    if not (valid_val(cc) and valid_val(tp) and valid_val(sn)):
        return HttpResponseBadRequest('Values must be between 0 and 20 or null')

    note = get_object_or_404(Note, id=note_id)
    ue = note.ue

    # permission: superuser or staff or instructor of the UE
    user = request.user
    def user_manages_ue(user, ue):
        return user.is_superuser or user.is_staff or ue.instructors.filter(pk=user.pk).exists() or (hasattr(user, 'ues') and user.ues.filter(pk=ue.pk).exists())
    if not user_manages_ue(user, ue):
        return HttpResponseForbidden()

    note.cc = cc
    note.tp = tp
    note.sn = sn
    note.save()

    return JsonResponse({'id': note.id, 'cc': note.cc, 'tp': note.tp, 'sn': note.sn, 'final': note.final, 'is_eliminated': note.is_eliminated})


@login_required
@require_POST
def note_create(request):
    # accept JSON body or form-encoded POST as fallback
    data = None
    if request.body:
        try:
            data = json.loads(request.body.decode())
        except Exception:
            data = None
    if data is None and request.POST:
        # build dict from POST fields
        data = {
            'etudiant_id': request.POST.get('etudiant_id'),
            'ue_id': request.POST.get('ue_id'),
            'cc': request.POST.get('cc'),
            'tp': request.POST.get('tp'),
            'sn': request.POST.get('sn'),
        }
        # convert numeric strings
        for k in ('etudiant_id', 'ue_id'):
            if data.get(k) is not None:
                try:
                    data[k] = int(data[k])
                except Exception:
                    pass
        for k in ('cc', 'tp', 'sn'):
            if data.get(k) is not None and data.get(k) != '':
                try:
                    data[k] = float(data[k])
                except Exception:
                    pass
            else:
                data[k] = None

    if not data:
        return HttpResponseBadRequest('Invalid JSON or empty POST body')

    etud_id = data.get('etudiant_id')
    ue_id = data.get('ue_id')
    cc = data.get('cc')
    tp = data.get('tp')
    sn = data.get('sn')

    if not etud_id or not ue_id:
        return HttpResponseBadRequest('etudiant_id and ue_id are required')

    # validate values
    def valid_val(x):
        return x is None or (isinstance(x, (int, float)) and 0 <= x <= 20)

    if not (valid_val(cc) and valid_val(tp) and valid_val(sn)):
        return HttpResponseBadRequest('Values must be between 0 and 20 or null')

    etud = get_object_or_404(Etudiant, id=etud_id)
    ue = get_object_or_404(UE, id=ue_id)

    # permission: superuser or staff or instructor of the UE can create
    user = request.user
    def user_manages_ue(user, ue):
        return user.is_superuser or user.is_staff or ue.instructors.filter(pk=user.pk).exists() or (hasattr(user, 'ues') and user.ues.filter(pk=ue.pk).exists())

    if not user_manages_ue(user, ue):
        return HttpResponseForbidden()

    note, created = Note.objects.get_or_create(etudiant=etud, ue=ue, defaults={'cc': cc, 'tp': tp, 'sn': sn})
    if not created:
        return HttpResponseBadRequest('Note already exists')

    return JsonResponse({'id': note.id, 'cc': note.cc, 'tp': note.tp, 'sn': note.sn, 'final': note.final, 'is_eliminated': note.is_eliminated})


# ---------- API pour cascade filters (département -> filière -> niveau -> ue) ----------
# public endpoints (GET)
def filieres_json(request):
    dep_id = request.GET.get('departement')
    if not dep_id:
        return JsonResponse({'filieres': []})
    filieres = Filiere.objects.filter(departement_id=dep_id).order_by('id')
    data = [{'id': f.id, 'nom': f.nom} for f in filieres]
    return JsonResponse({'filieres': data})


def niveaux_json(request):
    fil_id = request.GET.get('filiere')
    if not fil_id:
        return JsonResponse({'niveaux': []})
    niveaux = Niveau.objects.filter(etudiant__filiere_id=fil_id).distinct().order_by('id')
    data = [{'id': n.id, 'nom': n.nom} for n in niveaux]
    return JsonResponse({'niveaux': data})


def ues_json(request):
    fil_id = request.GET.get('filiere')
    niv_id = request.GET.get('niveau')
    if not fil_id or not niv_id:
        return JsonResponse({'ues': []})
    ues = UE.objects.filter(filiere_id=fil_id, niveau_id=niv_id).order_by('code')
    data = [{'id': u.id, 'nom': u.nom, 'code': u.code} for u in ues]
    return JsonResponse({'ues': data})


# ---------- Import notes from Excel ----------
@login_required
@require_POST
def notes_import_excel(request):
    """Import notes from Excel file. Expected columns: Nom, Matricule, CC, TP, SN"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    
    # Get UE ID from POST data
    ue_id = request.POST.get('ue_id')
    if not ue_id:
        return JsonResponse({'success': False, 'error': 'UE non spécifiée'})
    
    # Get uploaded file
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'Aucun fichier fourni'})
    
    file_obj = request.FILES['file']
    
    try:
        ue = UE.objects.get(pk=ue_id)
    except UE.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'UE introuvable'})
    
    # Parse Excel file
    try:
        wb = load_workbook(file_obj)
        ws = wb.active
        
        results = {
            'imported': 0,
            'updated': 0,
            'errors': []
        }
        
        # Iterate rows (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] and not row[1]:  # Skip empty rows
                continue
            
            nom = row[0]
            matricule = row[1]
            cc = row[2]
            tp = row[3]
            sn = row[4]
            
            # Validate that matricule exists
            try:
                etudiant = Etudiant.objects.get(matricule=matricule)
            except Etudiant.DoesNotExist:
                results['errors'].append(f"Row {row_idx}: Le matricule '{matricule}' de l'élève '{nom}' n'existe pas ou est incorrecte")
                continue
            
            # Convert and validate values
            try:
                cc = float(cc) if cc is not None else None
                tp = float(tp) if tp is not None else None
                sn = float(sn) if sn is not None else None
                
                # Validate ranges (0-20)
                for val, label in [(cc, 'CC'), (tp, 'TP'), (sn, 'SN')]:
                    if val is not None and (val < 0 or val > 20):
                        raise ValueError(f"{label} doit être entre 0 et 20")
            except (ValueError, TypeError) as e:
                results['errors'].append(f"Row {row_idx}: Erreur de conversion pour {matricule}: {str(e)}")
                continue
            
            # Create or update Note
            note, created = Note.objects.update_or_create(
                etudiant=etudiant,
                ue=ue,
                defaults={'cc': cc, 'tp': tp, 'sn': sn}
            )
            
            if created:
                results['imported'] += 1
            else:
                results['updated'] += 1
        
        return JsonResponse({
            'success': True,
            'imported': results['imported'],
            'updated': results['updated'],
            'errors': results['errors']
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erreur de lecture du fichier: {str(e)}'})


@login_required
def notes_export_excel(request):
    """Export notes to Excel file for a given UE and niveau."""
    if not request.user.is_staff:
        return HttpResponseForbidden()
    
    # Get UE and niveau IDs
    ue_id = request.GET.get('ue_id')
    niveau_id = request.GET.get('niveau_id')
    
    if not ue_id or not niveau_id:
        return HttpResponseBadRequest('UE and niveau required')
    
    try:
        ue = UE.objects.get(pk=ue_id)
        niveau = Niveau.objects.get(pk=niveau_id)
    except (UE.DoesNotExist, Niveau.DoesNotExist):
        return HttpResponseBadRequest('Invalid UE or niveau')
    
    # Get all students in this niveau with notes for this UE
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Notes'
    
    # Add headers
    headers = ['Nom', 'Matricule', 'CC', 'TP', 'SN']
    ws.append(headers)
    
    # Get students and their notes
    students = Etudiant.objects.filter(niveau=niveau).order_by('nom')
    notes_map = {n.etudiant_id: n for n in Note.objects.filter(ue=ue, etudiant__niveau=niveau)}
    
    # Add data rows
    for student in students:
        note = notes_map.get(student.id)
        ws.append([
            student.nom,
            student.matricule,
            note.cc if note and note.cc is not None else '',
            note.tp if note and note.tp is not None else '',
            note.sn if note and note.sn is not None else '',
        ])
    
    # Format header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as download
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"notes_{ue.code}_{niveau.nom.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------- Gestion des enseignants ----------
@login_required
def enseignants_list(request):
    if not request.user.is_staff:
        return HttpResponseForbidden()
    users = User.objects.all()
    return render(request, 'pages/enseignants_list_adminlte.html', {'users': users})


@login_required
def enseignant_create(request):
    if not request.user.is_staff:
        return HttpResponseForbidden()
    if request.method == 'POST':
        form = TeacherCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.email = form.cleaned_data.get('email')
            user.is_staff = form.cleaned_data.get('is_staff', False)
            user.save()
            messages.success(request, f"Enseignant {user.username} créé.")
            return redirect('enseignants_list')
    else:
        form = TeacherCreationForm()
    return render(request, 'pages/enseignant_form_adminlte.html', {'form': form})


@login_required
def enseignant_toggle_staff(request, user_id):
    if not request.user.is_staff:
        return HttpResponseForbidden()
    target = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        # only superuser can toggle staff status to avoid demoting main admin
        if not request.user.is_superuser:
            messages.error(request, "Seul le superuser peut modifier le statut staff.")
            return redirect('enseignants_list')
        if target == request.user:
            messages.error(request, "Vous ne pouvez pas modifier votre propre statut.")
            return redirect('enseignants_list')
        if target.is_superuser:
            messages.error(request, "Vous ne pouvez pas modifier le statut d'un superuser.")
            return redirect('enseignants_list')
        target.is_staff = not target.is_staff
        target.save()
        messages.success(request, f"{'Promu' if target.is_staff else 'Rétrogradé'} {target.username}.")
    return redirect('enseignants_list')


# ---------- API pour les filtres admin ----------
def etudiant_ues_json(request):
    """Return UEs for a given etudiant (for admin filtering)."""
    etudiant_id = request.GET.get('etudiant')
    if not etudiant_id:
        return JsonResponse([], safe=False)
    
    try:
        etudiant = Etudiant.objects.get(pk=etudiant_id)
        ues = UE.objects.filter(
            filiere=etudiant.filiere,
            niveau=etudiant.niveau
        ).values('id', 'code', 'nom').order_by('code')
        return JsonResponse(list(ues), safe=False)
    except (Etudiant.DoesNotExist, ValueError):
        return JsonResponse([], safe=False)

